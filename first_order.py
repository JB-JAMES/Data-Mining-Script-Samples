from core.models import CartItem, Order, CartItemStatusLog
from openpyxl import Workbook, load_workbook


def get_first_order_stat():

    order_cart_item_list = CartItemStatusLog.objects.filter(status=2).values_list('cart_item', flat=True)
    order_cart_items = CartItem.objects.filter(id__in=order_cart_item_list).prefetch_related(
        'member', 'order').order_by('order__deposit_date')
    print("order-cart-sucess")
    member_order_dict = {}
    first_order_list = []

    for item in order_cart_items:
        # 맴버별 first_order 배치시키기
        if member_order_dict.get(item.member_id) is None:
            member_order_dict[item.member_id] = item.order_id
            first_order_list.append(item.order_id)

    print("member-dic-sucess")

    # 7월
    month = ["2018-07-01", "2018-08-01", "2018-09-01", "2018-10-01"]
    sheet_title = ["7월", "8월", "9월"]
    idx = 0
    while month[idx] < "2018-10-01":
        first_orders = Order.objects.filter(deposit_date__gte=month[idx], deposit_date__lt=month[idx+1], id__in=first_order_list) \
            .prefetch_related('cart_items', 'cart_items__product')

        result = {}
        for order in first_orders:
            ck = 0
            id = 0
            product_ck = {}
            for item in order.cart_items.all():
                id = item.product_id

                if result.get(id) is None:
                    result[id] = {}
                    result[id]['name'] = item.product.name
                    result[id]['count'] = 0
                    result[id]['product count'] = 0
                    result[id]['total price'] = 0
                    result[id]['coupon'] = 0
                    result[id]['sale only'] = 0

                result[id]['product count'] += item.count
                result[id]['total price'] += item.price

                if product_ck.get(id) is None:
                    result[id]['count'] += 1
                    product_ck[id] = 1
                    ck += 1

                if order.use_coupon_price == 10000:
                    result[id]['coupon'] += 1
            if ck == 1:
                result[id]['sale only'] += 1

        sorted_ids = sorted(result, key=lambda x: result[x]["count"], reverse=True)

        try:
            wb = load_workbook('FO.xlsx')
        except:
            wb = Workbook()

        ws = wb.create_sheet(sheet_title[idx])

        _ = ws.cell(column=1, row=1, value='제품번호')
        _ = ws.cell(column=2, row=1, value='제품명')
        _ = ws.cell(column=3, row=1, value='구매횟수')
        _ = ws.cell(column=4, row=1, value='구매수량')
        _ = ws.cell(column=5, row=1, value='거래액')
        _ = ws.cell(column=6, row=1, value='단독 구매횟수')
        _ = ws.cell(column=7, row=1, value='쿠폰 사용횟수')

        row = 2
        for id in sorted_ids:
            _ = ws.cell(column=1, row=row, value=id)
            _ = ws.cell(column=2, row=row, value=result[id]['name'])
            _ = ws.cell(column=3, row=row, value=result[id]['count'])
            _ = ws.cell(column=4, row=row, value=result[id]['product count'])
            _ = ws.cell(column=5, row=row, value=result[id]['total price'])
            _ = ws.cell(column=6, row=row, value=result[id]['sale only'])
            _ = ws.cell(column=7, row=row, value=result[id]['coupon'])
            row += 1

        wb.save('FO.xlsx')
        idx += 1